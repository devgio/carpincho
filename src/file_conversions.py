import openpyxl, os, csv, xlrd
from dateutil.parser import parse as parse_date
import xml.etree.ElementTree as ET

def xml2d_to_xlsx(xml_file, new_row_tag='line', value_tag='data', remove_source=True):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Sheet1'
	tree = ET.parse(xml_file)
	row = 0
	for element in tree.iter():
		if element.tag == new_row_tag:
			col = 1
			row += 1
		elif element.tag == value_tag:
			try:
				element.text = float(element.text)
			except:
				pass
			ws.cell(row=row, column=col).value = element.text
			col += 1
	save_file = os.path.splitext(xml_file)[0]+'.xlsx'
	wb.save(save_file)
	if remove_source:
		os.remove(xml_file)
	return save_file


def csv_to_xlsx(file, template=False, floats=range(5,13), dates=range(0,1)):
    '''
    floats: A list of column indexes in which floats appear
    dates: a list of column indexes in which floats appear
    '''
    folder = os.path.split(file)[0]
    file_name = os.path.split(file)[-1].split('.')[0]
    wb = openpyxl.Workbook()
    if template:
        dir_path = os.path.dirname(os.path.realpath(__file__))
        wb = openpyxl.load_workbook(os.path.join(dir_path, template))
    ws = wb.active
    with open(file, 'rt') as f:
        csv.register_dialect('commas', delimiter=',')
        reader = csv.reader(f, dialect='commas')
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                try:
                    if column_index in floats:
                        cell = float(cell)
                    elif column_index in dates:
                        cell = parse_date(cell).date()
                except:
                    pass
                ws.cell(column=column_index+1, row=row_index+1).value = cell
    excel_file = os.path.join(folder, f'{file_name}.xlsx')
    wb.save(excel_file)
    os.remove(file)
    return excel_file


def xls_to_xlsx(xls_file, delete_source=True):
	directory, file = os.path.split(xls_file)
	if file.split('.')[-1].lower() != 'xls':
		return False
	xls_wb = xlrd.open_workbook(xls_file)
	wb = openpyxl.Workbook()
	del wb['Sheet']
	for xls_ws in xls_wb.sheets():
		wb.create_sheet(xls_ws.name)
		ws = wb[xls_ws.name]
		for r in range(xls_ws.nrows):
			for c in range(xls_ws.ncols):
				ws.cell(row=r+1, column=c+1).value = xls_ws.cell(r,c).value
	xlsx_file = os.path.join(directory, file + 'x')
	wb.save(xlsx_file)
	if delete_source:
		os.remove(xls_file)
	return xlsx_file
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_interval, get_column_letter, coordinate_from_string, column_index_from_string


def df_from_range(ws, range_string):
	col_start, col_end = re.findall("[A-Z]+", range_string)
	data_rows = []
	for row in ws[range_string]:
		data_rows.append([cell.value for cell in row])
	return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


def search(ws, keyword, ocurrence=1, exact_match=False):
	for row_number, row in enumerate(ws):
		for col_number, cell in enumerate(row):
			if cell.value and keyword in str(cell.value):
				if exact_match and cell.value != keyword:
					continue
				ocurrence -= 1
				if ocurrence == 0:
					return (col_number+1, row_number+1)


def df_from_search(ws, 
	top_left_keyword, top_left_exact_match=False, top_left_col_offset=0, top_left_row_offset=0, top_left_occurrence=1,
	bottom_right_keyword=None, bottom_right_exact_match=False, bottom_right_col_offset=0, bottom_right_row_offset=0, bottom_right_occurrence=1):

	top_left_offset = (top_left_col_offset, top_left_row_offset)
	top_left = search(ws, top_left_keyword, top_left_occurrence, top_left_exact_match)
	top_left = tuple(sum(_) for _ in zip(top_left, top_left_offset))

	if not bottom_right_keyword:
		bottom_right_keyword = top_left_keyword
	bottom_right_offset = (bottom_right_col_offset, bottom_right_row_offset)
	bottom_right = search(ws, bottom_right_keyword, bottom_right_occurrence, bottom_right_exact_match)
	bottom_right = tuple(sum(_) for _ in zip(bottom_right, bottom_right_offset))

	range_string = f'{get_column_letter(top_left[0])}{top_left[1]}:{get_column_letter(bottom_right[0])}{bottom_right[1]}'
	
	return df_from_range(ws, range_string)


def paste_df(df, ws, top_left_cell):
	top_left_col, top_left_row = coordinate_from_string(top_left_cell)
	top_left_col = column_index_from_string(top_left_col)
	rows, cols = df.shape
	for r in range(rows):
		for c in range(cols):
			ws.cell(row=top_left_row+r, column=top_left_col+c).value = df.iat[r,c]


def pull_date(file, cell_range, worksheet=0, max_or_min=max, dayfirst=True):
	import xlrd, datetime, os
	import dateutil.parser as date_parser
	
	top_left_cell, bottom_right_cell = (cell.upper() for cell in cell_range.split(':'))
	top_left_col, top_left_row = coordinate_from_string(top_left_cell)
	top_left_col = column_index_from_string(top_left_col)
	bottom_right_col, bottom_right_row = coordinate_from_string(bottom_right_cell)
	bottom_right_col = column_index_from_string(bottom_right_col)

	directory, file_name = os.path.split(file)
	file_name, file_type = os.path.splitext(file_name)

	if file_type.lower() == '.xls':
		wb = xlrd.open_workbook(file, logfile=open(os.devnull, 'w')) # Added the logfile=open(os.devnull, 'w') to supress warnings.
		ws = wb.sheets()[worksheet]
		top_left_col -= 1
		bottom_right_col -= 1
		top_left_row -= 1
		bottom_right_row -= 1
		max_col, max_row = ws.ncols-1, ws.nrows-1
	if file_type.lower() in ('.xlsx', '.xlsm'):
		wb = load_workbook(file)
		ws = wb.worksheets[worksheet]
		max_col, max_row = ws.max_column, ws.max_row


	dates = set()
	for col in range(top_left_col, bottom_right_col+1):
		if col <= max_col:
			for row in range(top_left_row, bottom_right_row+1):
				if row <= max_row:
					if file_type.lower() in ('.xlsx', '.xlsm'):
						cell_value = ws.cell(row=row, column=col).value
					else:
						cell_value = ws.cell(row, col).value
					date = None
					if file_type.lower() == '.xls':
						try:
							date = datetime.datetime(*xlrd.xldate_as_tuple(cell_value, wb.datemode))
						except: pass
					if not date:
						try:
							if type(cell_value) == datetime.datetime:
								date = cell_value
							else:
								date = date_parser.parse(cell_value, dayfirst=dayfirst, fuzzy=True)
						except: pass
					if type(date) == datetime.datetime:
						dates.add(date)
	try:
		return max_or_min(dates)
	except:
		return None
